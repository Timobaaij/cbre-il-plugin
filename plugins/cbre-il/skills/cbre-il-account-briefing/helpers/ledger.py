"""Source Ledger builder, validator, merger, and exporter.

The Source Ledger is the spine of the deck's defensibility: every material claim
maps to one row with a retrievable source. This module enforces the column
schema and Claim-ID uniqueness so the ledger cannot quietly drift, merges the
five research agents' candidate source files into one canonical ledger, and
exports it to .xlsx for delivery.

Canonical columns (see reference/deliverables-and-ledger.md):
  claim_id, slide, claim, source_url, source_tier,
  figure_or_quote_at_source, source_pub_date, retrieved_date,
  confidence_band, binding_test, verified

Required-at-merge (a row missing any of these is rejected):
  claim_id, claim, source_url, source_tier, retrieved_date

CLI:
  python ledger.py merge  <out.csv> <agentA.sources.csv> [agentB.sources.csv ...]
  python ledger.py validate <ledger.csv>
  python ledger.py export <ledger.csv> <out.xlsx>
  python ledger.py backfill-slides <content_plan.json> <ledger.csv>
"""
from __future__ import annotations

import argparse
import csv
import json
import sys
from pathlib import Path
from typing import List
from urllib.parse import urlparse

COLUMNS = [
    "claim_id", "slide", "claim", "source_url", "source_tier",
    "figure_or_quote_at_source", "source_pub_date", "retrieved_date",
    "confidence_band", "binding_test", "verified",
]
REQUIRED_AT_MERGE = ["claim_id", "claim", "source_url", "source_tier", "retrieved_date"]
VALID_TIERS = {"1", "2", "3", "4", "5", "6"}


class LedgerError(Exception):
    pass


def _read_rows(path: str) -> List[dict]:
    with open(path, newline="", encoding="utf-8-sig") as f:
        return list(csv.DictReader(f))


def _norm(row: dict) -> dict:
    """Return a row with every canonical column present (missing -> '')."""
    return {c: (row.get(c) or "").strip() for c in COLUMNS}


def reference_rows(ledger_path) -> List[dict]:
    """The ONE canonical, de-duplicated, numbered reference list, used by BOTH the
    rendered references slide (build_deck.r_references) and build_report.json
    (build_deck._reference_map), and counted by gate_runner.reconcile. Keeping a
    single source of numbering means the [N] markers on the deck, the report's
    numbering, and reconcile's expected marker count always agree.

    A reference line is one distinct (host, tier, pub_date); the first
    cited+verified+slide-mapped ledger row for that key owns the number. Rows that
    are struck (verified=no) or not mapped to a slide are excluded."""
    out: List[dict] = []
    seen = set()
    if not (ledger_path and Path(ledger_path).exists()):
        return out
    for r in _read_rows(str(ledger_path)):
        if (r.get("verified") or "").lower() == "no" or not (r.get("slide") or "").strip():
            continue
        host = urlparse(r.get("source_url", "")).netloc.replace("www.", "")
        tier = (r.get("source_tier") or "").strip()
        pub = (r.get("source_pub_date") or "").strip()
        key = (host, tier, pub)
        if key in seen:
            continue
        seen.add(key)
        out.append({"n": len(out) + 1, "host": host, "tier": tier,
                    "pub_date": pub, "claim_id": (r.get("claim_id") or "").strip()})
    return out


def backfill_slides(plan_path: str, ledger_path: str) -> dict:
    """Derive the ledger 'slide' column from the (frozen) content plan, IN PLACE.

    Each cited claim_id's slide is the slide_no of the FIRST slide (deck order) that
    cites it, looking at the slide-level claim_ids and every cell-level claim_ids
    (first wins). Struck rows (verified=no) and uncited verified rows are left blank,
    so they never appear on the references slide. This is exactly the derivation the
    spec means by "the generator back-fills it" (reference/content-plan-spec.md) and
    that qa1 demands.

    Idempotent: the column is recomputed from the plan on every run (the prior 'slide'
    value is ignored), so it is safe to re-run after any plan edit. Preserves the
    canonical COLUMNS order and writes utf-8. Also repairs a short-header ledger via
    _norm. Returns {filled, total, cited, missing} for the caller to print, where
    'missing' is plan-cited claim_ids absent from the ledger."""
    plan = json.loads(Path(plan_path).read_text(encoding="utf-8"))
    first: dict = {}
    for s in plan.get("slides", []) or []:
        no = s.get("slide_no")
        if no is None or str(no).strip() == "":
            continue
        cids = list(s.get("claim_ids", []) or [])
        for r in s.get("scene", []) or []:
            for c in r.get("cells", []) or []:
                cids += list(c.get("claim_ids", []) or [])
        for cid in cids:
            if cid:
                first.setdefault(cid, no)
    rows = [_norm(r) for r in _read_rows(ledger_path)]
    filled = 0
    for row in rows:
        cid = row["claim_id"]
        if (row.get("verified") or "").lower() == "no":
            row["slide"] = ""           # struck -> never on a slide
        elif cid in first:
            row["slide"] = str(first[cid])
            filled += 1
        else:
            row["slide"] = ""           # uncited verified -> not on references
    with open(ledger_path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=COLUMNS)
        w.writeheader()
        for row in rows:
            w.writerow(row)
    ledger_ids = {r["claim_id"] for r in rows}
    missing = sorted(cid for cid in first if cid not in ledger_ids)
    return {"filled": filled, "total": len(rows), "cited": len(first), "missing": missing}


def validate(path: str, *, at_merge: bool = False) -> List[str]:
    """Return a list of problems. Empty list means the ledger is clean."""
    problems: List[str] = []
    rows = _read_rows(path)
    seen = {}
    for i, raw in enumerate(rows, start=2):  # header is line 1
        row = _norm(raw)
        cid = row["claim_id"]
        if not cid:
            problems.append(f"line {i}: missing claim_id")
            continue
        if cid in seen:
            problems.append(f"line {i}: duplicate claim_id '{cid}' (first seen line {seen[cid]})")
        seen[cid] = i
        for col in REQUIRED_AT_MERGE:
            if not row[col]:
                problems.append(f"line {i} ({cid}): missing required field '{col}'")
        if row["source_tier"] and row["source_tier"] not in VALID_TIERS:
            problems.append(f"line {i} ({cid}): source_tier '{row['source_tier']}' not in 1-6")
        if not at_merge:
            if not row["confidence_band"]:
                problems.append(f"line {i} ({cid}): missing confidence_band")
            if row["verified"] and row["verified"].lower() not in ("yes", "no"):
                problems.append(f"line {i} ({cid}): verified must be yes/no, got '{row['verified']}'")
    return problems


def _dedupe_key(row: dict) -> tuple:
    # Two candidate records are "the same source claim" only if URL, figure, AND
    # claim text all match. Keying on URL + figure alone would collapse two
    # genuinely different claims that share one source URL (e.g. an annual-report
    # PDF) and happen to cite the same round figure (two facilities both
    # "50,000 sqm"); including the claim text prevents that data loss.
    return (row["source_url"].lower(),
            row["figure_or_quote_at_source"].lower().strip(),
            row["claim"].lower().strip())


def merge(out_path: str, source_paths: List[str]) -> dict:
    """Merge agent *.sources.csv files into one canonical ledger.

    De-dupes identical-source claims (same URL + same figure/quote), keeping the
    first (highest-tier wins on ties), assigns stable Claim IDs (agent prefix
    preserved), and REJECTS any row missing a required field.
    """
    kept: List[dict] = []
    by_dedupe = {}
    rejected: List[str] = []
    used_ids = set()

    for sp in source_paths:
        for i, raw in enumerate(_read_rows(sp), start=2):
            row = _norm(raw)
            missing = [c for c in REQUIRED_AT_MERGE if not row[c]]
            if missing:
                rejected.append(f"{Path(sp).name} line {i}: missing {missing} -> rejected")
                continue
            if row["source_tier"] not in VALID_TIERS:
                rejected.append(f"{Path(sp).name} line {i}: bad tier '{row['source_tier']}' -> rejected")
                continue
            key = _dedupe_key(row)
            if key in by_dedupe:
                # Keep the higher-tier (lower number) record.
                existing = by_dedupe[key]
                if int(row["source_tier"]) < int(existing["source_tier"]):
                    kept[kept.index(existing)] = row
                    by_dedupe[key] = row
                continue
            # Ensure a unique claim_id.
            cid = row["claim_id"]
            base = cid
            n = 1
            while cid in used_ids:
                n += 1
                cid = f"{base}-{n}"
            row["claim_id"] = cid
            used_ids.add(cid)
            by_dedupe[key] = row
            kept.append(row)

    with open(out_path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=COLUMNS)
        w.writeheader()
        for row in kept:
            w.writerow(row)

    return {"kept": len(kept), "rejected": rejected, "out": out_path}


def export(ledger_path: str, xlsx_path: str) -> str:
    """Export the ledger to a filterable .xlsx. Falls back to .csv copy if
    openpyxl is unavailable (with a warning), so the pipeline never hard-fails."""
    rows = [_norm(r) for r in _read_rows(ledger_path)]
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        alt = str(Path(xlsx_path).with_suffix(".csv"))
        with open(alt, "w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=COLUMNS)
            w.writeheader()
            w.writerows(rows)
        sys.stderr.write(f"WARNING: openpyxl not installed; wrote {alt} instead of xlsx\n")
        return alt

    wb = Workbook()
    ws = wb.active
    ws.title = "Source Ledger"
    header_fill = PatternFill("solid", fgColor="1F3864")
    header_font = Font(color="FFFFFF", bold=True)
    ws.append(COLUMNS)
    for c in range(1, len(COLUMNS) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center")
    for row in rows:
        ws.append([row[c] for c in COLUMNS])
    # Widths + freeze + autofilter.
    widths = {"claim": 48, "source_url": 40, "figure_or_quote_at_source": 36, "claim_id": 12}
    for idx, col in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = widths.get(col, 16)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{ws.max_row}"
    wb.save(xlsx_path)
    return xlsx_path


def main() -> None:
    p = argparse.ArgumentParser(description="Source Ledger build/validate/merge/export.")
    sub = p.add_subparsers(dest="cmd", required=True)

    m = sub.add_parser("merge")
    m.add_argument("out")
    m.add_argument("sources", nargs="+")

    v = sub.add_parser("validate")
    v.add_argument("ledger")
    v.add_argument("--at-merge", action="store_true")

    e = sub.add_parser("export")
    e.add_argument("ledger")
    e.add_argument("out")

    b = sub.add_parser("backfill-slides")
    b.add_argument("plan")
    b.add_argument("ledger")

    a = p.parse_args()
    if a.cmd == "merge":
        res = merge(a.out, a.sources)
        print(f"merged {res['kept']} rows -> {res['out']}")
        for r in res["rejected"]:
            print("  REJECTED:", r)
        probs = validate(a.out, at_merge=True)
        if probs:
            print("POST-MERGE VALIDATION PROBLEMS:")
            for pr in probs:
                print("  -", pr)
            sys.exit(1)
    elif a.cmd == "validate":
        probs = validate(a.ledger, at_merge=a.at_merge)
        if probs:
            print(f"INVALID ({len(probs)} problems):")
            for pr in probs:
                print("  -", pr)
            sys.exit(1)
        print("OK: ledger is valid")
    elif a.cmd == "export":
        out = export(a.ledger, a.out)
        print(f"exported -> {out}")
    elif a.cmd == "backfill-slides":
        res = backfill_slides(a.plan, a.ledger)
        print(f"back-filled slide on {res['filled']} of {res['total']} ledger rows "
              f"({res['cited']} distinct claims cited in the plan)")
        for cid in res["missing"]:
            print("  WARNING: plan cites claim_id absent from ledger:", cid)


if __name__ == "__main__":
    main()
